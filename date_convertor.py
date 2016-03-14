from openpyxl import *
from datetime import *

wb_in = load_workbook('входящий.xlsx')
wb_out = load_workbook('конечный.xlsx')


def convert_d(in_wb, out_wb, d):
    ws_in = wb_in.active
    ws_out = wb_out.active

    for i in range(1, 71):
        a = ws_in.cell(row=i, column=d)
        b = a.value.strftime('%d.%m.%Y %H:%M')
        ws_out.cell(row=i, column=d).value = b

convert_d(wb_in, wb_out, 4)
convert_d(wb_in, wb_out, 5)

wb_out.save("savesmth.xlsx")
